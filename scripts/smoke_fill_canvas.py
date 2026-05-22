#!/usr/bin/env python3
"""Smoke-test for skills/sharebird-unicorn-canvas/scripts/fill_canvas.py.

Runs the fill script against a deliberately-non-cybersecurity example
(health-tech) so we exercise the bundled template with a payload the
worked-example data doesn't already cover.
"""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
FILL_SCRIPT = REPO_ROOT / "skills" / "sharebird-unicorn-canvas" / "scripts" / "fill_canvas.py"

SAMPLE_PAYLOAD = {
    "brand_line": "We give every clinician the same patient story before they walk in the room.",
    "pains": [
        {
            "eyebrow": "PAIN 1 · DOMINANT",
            "title": "Specialists are starting visits without the primary-care context they need.",
        },
        {
            "eyebrow": "PAIN 2",
            "title": "Care teams repeat the same intake questions across every visit.",
        },
        {
            "eyebrow": "PAIN 3",
            "title": "Risk-stratified patients fall through the gaps between visits.",
        },
    ],
    "differentiation": [
        "One patient summary that auto-syncs to the visit type 90 seconds before the appointment.",
        "Intake questions disappear once any teammate has answered them anywhere in the system.",
        "Risk flags surface in the clinician's existing inbox — not in a separate dashboard.",
    ],
    "outcomes": [
        "Cut specialist visit prep from 12 minutes to under 2 at Northwell.",
        "Reduced duplicate intake by 71% across 14 clinics in the Mass General pilot.",
        "Caught 38 high-risk patients in 90 days that the prior workflow missed.",
    ],
    "modules": [
        ["Visit Brief Generator", "Care Plan Sync", "Specialist Handoff Pack"],
        ["Universal Intake", "Patient Profile Graph", "Form-suppression API"],
        ["Risk Inbox", "Care-Gap Alerts", "Population Cohorts"],
    ],
    "power_plays": [
        "Target health-system specialty groups with >20% no-context-start rate; trigger on a CMIO appointment; lead with the 90-second visit brief demo; aim for a 30-day pilot on one specialty line.",
        "Target multi-clinic primary-care groups doing >5 form revisions per quarter; trigger on a new EHR optimization committee charter; lead with the intake-suppression ROI calculator; aim for a workflow audit meeting.",
        "Target ACOs with risk-bearing contracts; trigger on a population-health director hire; lead with the Mass General case study; aim for a quarterly business review.",
    ],
}


def main() -> int:
    if not FILL_SCRIPT.exists():
        print(f"Missing fill script: {FILL_SCRIPT}", file=sys.stderr)
        return 1

    with tempfile.TemporaryDirectory() as tmpdir:
        payload_path = Path(tmpdir) / "payload.json"
        out_path = Path(tmpdir) / "smoke-output.xlsx"
        payload = dict(SAMPLE_PAYLOAD)
        payload["output_path"] = str(out_path)
        payload_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

        result = subprocess.run(
            [sys.executable, str(FILL_SCRIPT), "--input", str(payload_path)],
            capture_output=True,
            text=True,
            check=False,
        )

        if result.returncode != 0:
            print("fill_canvas.py exited non-zero", file=sys.stderr)
            print(result.stderr, file=sys.stderr)
            return result.returncode

        if not out_path.exists():
            print(f"Expected output not written: {out_path}", file=sys.stderr)
            return 1

        size = out_path.stat().st_size
        if size < 5_000:
            print(f"Output suspiciously small: {size} bytes", file=sys.stderr)
            return 1

        print(f"OK — wrote {size} bytes to {out_path}")
        # Re-open with openpyxl to confirm we can round-trip
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("openpyxl not installed; install with `pip install openpyxl`", file=sys.stderr)
            return 1
        workbook = load_workbook(out_path)
        assert "Worked example" in workbook.sheetnames, "Worked example tab missing"
        sheet = workbook["Worked example"]
        assert sheet["B5"].value == SAMPLE_PAYLOAD["brand_line"], "Brand line not written"
        print("OK — bundled template + Worked example tab readable; brand line round-tripped")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
