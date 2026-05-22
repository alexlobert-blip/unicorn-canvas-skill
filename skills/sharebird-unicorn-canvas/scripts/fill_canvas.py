#!/usr/bin/env python3
"""Fill a copy of the bundled Unicorn Canvas template with user answers.

Reads a JSON payload (stdin or --input <file>), writes the answers into the
``Worked example`` tab of a fresh copy of ``templates/canvas-blank.xlsx``,
and saves the result to ``output_path`` (defaults to
``~/Downloads/unicorn-canvas.xlsx``).

The cell addresses below mirror the layout produced by the upstream
``build_sharebird_templates.py`` builder — keep them in sync with that file.

Payload shape::

    {
      "brand_line": "…",
      "pains": [
        {"eyebrow": "PAIN 1 · DOMINANT", "title": "…"},
        {"eyebrow": "PAIN 2", "title": "…"},
        {"eyebrow": "PAIN 3", "title": "…"}
      ],
      "differentiation": ["…", "…", "…"],
      "outcomes": ["…", "…", "…"],
      "modules": [
        ["…", "…", "…"],
        ["…", "…", "…"],
        ["…", "…", "…"]
      ],
      "power_plays": ["…", "…", "…"],
      "output_path": "~/Downloads/unicorn-canvas-acme.xlsx"
    }
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - import-time guard
    sys.stderr.write(
        "openpyxl is required. Install with `pip install openpyxl`.\n"
    )
    raise SystemExit(1) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = SCRIPT_DIR.parent / "templates" / "canvas-blank.xlsx"
WORKED_EXAMPLE_TAB = "Worked example"

# Cell addresses verified against canvas-blank.xlsx (Worked example tab).
# Pain headers are stored as ``"EYEBROW | title"`` in a single cell.
BRAND_LINE_CELL = "B5"

PAIN_HEADER_CELLS = ("B7", "C7", "D7")

DIFFERENTIATION_CELLS = ("B9", "C9", "D9")
OUTCOME_CELLS = ("B11", "C11", "D11")

# Modules are three per pain, rows 12 / 13 / 14.
MODULE_ROWS = (12, 13, 14)
MODULE_COLS = ("B", "C", "D")

POWER_PLAY_CELLS = ("B15", "C15", "D15")


def _load_payload(path: str | None) -> dict[str, Any]:
    if path:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    return json.load(sys.stdin)


def _resolve_output(path_str: str | None) -> Path:
    if not path_str:
        return Path.home() / "Downloads" / "unicorn-canvas.xlsx"
    expanded = Path(path_str).expanduser()
    if expanded.is_dir():
        return expanded / "unicorn-canvas.xlsx"
    return expanded


def fill_canvas(payload: dict[str, Any], output_path: Path) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Template missing: {TEMPLATE_PATH}. Re-install the plugin or rebuild templates."
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, output_path)

    workbook = load_workbook(output_path)
    if WORKED_EXAMPLE_TAB not in workbook.sheetnames:
        raise ValueError(
            f"Expected tab '{WORKED_EXAMPLE_TAB}' missing in {output_path}"
        )
    sheet = workbook[WORKED_EXAMPLE_TAB]

    brand_line = payload.get("brand_line", "")
    if brand_line:
        sheet[BRAND_LINE_CELL] = brand_line

    pains = payload.get("pains", []) or []
    for idx, pain in enumerate(pains[:3]):
        if isinstance(pain, dict):
            eyebrow = pain.get("eyebrow", "").strip()
            title = pain.get("title", "").strip()
        else:
            eyebrow = ""
            title = str(pain).strip()
        combined = " | ".join(part for part in (eyebrow, title) if part)
        if combined:
            sheet[PAIN_HEADER_CELLS[idx]] = combined

    for idx, value in enumerate((payload.get("differentiation") or [])[:3]):
        if value:
            sheet[DIFFERENTIATION_CELLS[idx]] = value

    for idx, value in enumerate((payload.get("outcomes") or [])[:3]):
        if value:
            sheet[OUTCOME_CELLS[idx]] = value

    modules = payload.get("modules") or []
    for col_idx, column in enumerate(modules[:3]):
        if not isinstance(column, list):
            continue
        for row_idx, value in enumerate(column[:3]):
            if value:
                sheet[f"{MODULE_COLS[col_idx]}{MODULE_ROWS[row_idx]}"] = value

    for idx, value in enumerate((payload.get("power_plays") or [])[:3]):
        if value:
            sheet[POWER_PLAY_CELLS[idx]] = value

    workbook.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        help="Path to a JSON file with the payload. Reads stdin if omitted.",
    )
    parser.add_argument(
        "--output",
        help="Override the output path. Defaults to payload.output_path or ~/Downloads/unicorn-canvas.xlsx.",
    )
    args = parser.parse_args()

    try:
        payload = _load_payload(args.input)
    except json.JSONDecodeError as exc:
        sys.stderr.write(f"Invalid JSON payload: {exc}\n")
        return 2

    output_path = _resolve_output(args.output or payload.get("output_path"))

    try:
        written = fill_canvas(payload, output_path)
    except (FileNotFoundError, ValueError) as exc:
        sys.stderr.write(f"{exc}\n")
        return 1

    print(str(written))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
